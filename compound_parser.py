from dataclasses import dataclass
from csv import DictWriter  
from openpyxl import load_workbook
import logging

@dataclass
class Compound:
    id: str
    pct_inhibition: float
    pf_gametocyte_ic50_avg: float
    pf_gametocyte_ic50_sd: float
    hepg2_cytotoxicity_ic50: str # ">25" is an option
    hepg2_pf_ic50_ratio: str
    pc_asexual_ic50: float


def parse_inhibition_xlsx_file(filepath):
    """
    Parses the inhibition file, which follows the repeating column format:

    <unused> | Compound ID | Inhibition | ...

    Args:
        filepath (str): Path to the xlsx file.
    
    Returns:    
        inhibition_dict: A dict with compound ID as keys, and inhibition values as values.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file does not follow the expected format.
    """
    inhibition_dict = {}
    wb = load_workbook(filepath)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    for row in rows:
        if not any(row): # skip empty lines
            continue
        for i, cell in enumerate(row):
            if isinstance(cell, str) and cell.startswith("TCMDC"):
                inhibition_dict[cell] = row[i+1] # grab the inhib data immediately to the right of the compound ID
    else:
        logging.debug(f"Parsed {len(inhibition_dict)} inhibition values.")

    if not inhibition_dict:
        logging.warning("No inhibition values were parsed.")
    return inhibition_dict

def parse_ic50_file(filepath, inhibition_dict):
    """
    Parses the IC50 XLSX file, which has TCMDC IDs in the first column of each row, followed by their IC50 data. 
    Args:
        filepath (str): Path to the xlsx file.
        inhibition_dict (dict): A dict with compound ID as keys, and inhibition values as values.
    
    Returns:
        compound_list: A list of Compound objects.
    """
    compound_list = []
    wb = load_workbook(filepath)
    ws = wb.active
    values = ws.iter_rows(values_only=True)
    next(values) # skip header rows
    next(values) 
    for value in values:
        if not isinstance(value[0], str) and value[0].startswith("TCMDC"):
            continue
        try:
            inhibition_value = inhibition_dict[value[0]]
        except KeyError as e:
            logging.error(f"Could not find inhibition value for {value[0]}")
            raise e
        compound_list.append(create_compound(value, inhibition_value))
    else:
        logging.debug(f"Parsed IC50 data for {len(compound_list)} compounds.")
    return compound_list

def create_compound(value, inhibition_value):
    try:
        return Compound(
            id=value[0],
            pct_inhibition=inhibition_value,
            pf_gametocyte_ic50_avg=value[3],
            pf_gametocyte_ic50_sd=value[4],
            hepg2_cytotoxicity_ic50=value[6],
            hepg2_pf_ic50_ratio=value[8],
            pc_asexual_ic50=value[9])
    except IndexError as e:
        logging.error(f"Could not find IC50 values for {value[0]}")
        raise e



def output_csv(compound_list, output_filepath):
    """
    Outputs the compound list as a CSV file.

    Args:
        compound_list (list): A list of Compound objects.
        output_filepath (str): Path to the output CSV file.
    """
    with open(output_filepath, "w") as f:
        writer = DictWriter(f, fieldnames=Compound.__dataclass_fields__.keys())
        try:
            writer.writeheader()
            for compound in compound_list:
                writer.writerow(compound.__dict__)
            else:
                logging.info(f"Wrote info for {len(compound_list)} compounds to {output_filepath}.")
        except Exception as e:
            logging.error(f"Could not write to {output_filepath}.")
            raise e

def main():
    logging.basicConfig(level=logging.INFO)
    logging.info("Starting...")
    inhibition_dict = parse_inhibition_xlsx_file("data/inhibition.xlsx")
    compound_list = parse_ic50_file("data/ic50.xlsx", inhibition_dict)
    output_csv(compound_list, "output.csv")
    logging.info("Done.")

if __name__ == "__main__":
    main()